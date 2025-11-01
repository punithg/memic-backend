"""
File conversion utilities for converting documents to PDF using LibreOffice.

This module is self-contained for easy extraction to a separate microservice.
Ported from rag-file-conversion Azure Function with simplifications.
"""
import os
import time
import subprocess
import tempfile
import shutil
import logging
from typing import Optional, Tuple
from pathlib import Path

from app.config import settings

logger = logging.getLogger(__name__)

# Lazy import for openpyxl - only imported when actually needed
# This allows the app to start even if openpyxl is not installed
def _import_openpyxl():
    """Lazy import openpyxl only when needed."""
    try:
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        return openpyxl, Font, get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel file conversion. "
            "Install it with: pip install openpyxl==3.1.2"
        )


def needs_conversion(filename: str) -> bool:
    """
    Check if file needs LibreOffice conversion to PDF.
    
    Args:
        filename: Name of the file to check
        
    Returns:
        True if file needs conversion, False otherwise
    """
    filename_lower = filename.lower()
    
    # Skip - go directly to parser (no conversion)
    if filename_lower.endswith(('.pdf', '.json')):
        return False
    
    # Skip - parser handles these formats directly
    if filename_lower.endswith(('.xlsx', '.pptx')):
        return False
        
    # Skip - parser handles audio transcription
    if filename_lower.endswith(('.mp3', '.wav', '.m4a', '.flac', '.ogg', '.aac')):
        return False
        
    # Skip - parser handles email conversion
    if filename_lower.endswith(('.eml', '.msg')):
        return False
    
    # Convert these formats to PDF
    if filename_lower.endswith(('.doc', '.docx', '.xls', '.ppt')):
        return True
        
    # Convert images to PDF
    if filename_lower.endswith(('.jpg', '.jpeg', '.png')):
        return True
    
    # Unknown format - attempt conversion
    logger.warning(f"Unknown file format: {filename}. Will attempt conversion.")
    return True


def create_temp_directory() -> str:
    """
    Create a temporary directory for file processing.
    
    Returns:
        Path to the created temporary directory
    """
    try:
        # Create unique temp directory using timestamp
        epoch_time = time.time()
        epoch_time_str = str(epoch_time).replace(".", "_")
        temp_dir = os.path.join(tempfile.gettempdir(), f"memic_conversion_{epoch_time_str}")
        
        os.makedirs(temp_dir, exist_ok=True)
        
        if not os.path.exists(temp_dir):
            raise RuntimeError(f"Failed to create temporary directory: {temp_dir}")
        if not os.access(temp_dir, os.W_OK):
            raise RuntimeError(f"Temporary directory is not writable: {temp_dir}")
            
        logger.info(f"Created temporary directory: {temp_dir}")
        return temp_dir
        
    except Exception as e:
        logger.error(f"Error creating temporary directory: {str(e)}")
        raise


def cleanup_temp_files(temp_dir: str) -> None:
    """
    Clean up temporary files and directory.
    
    Args:
        temp_dir: Path to temporary directory to clean up
    """
    if temp_dir and os.path.exists(temp_dir):
        try:
            logger.info(f"Starting cleanup of temporary directory: {temp_dir}")
            shutil.rmtree(temp_dir)
            logger.info("Cleanup completed successfully")
        except Exception as e:
            logger.warning(f"Error cleaning up temporary files: {str(e)}")


def preprocess_excel(input_path: str, output_path: str) -> str:
    """
    Preprocess Excel file before PDF conversion.
    Adds sheet titles and formats cells for better PDF output.
    
    Args:
        input_path: Path to input Excel file
        output_path: Path where preprocessed file should be saved
        
    Returns:
        Path to the preprocessed Excel file
    """
    try:
        # Lazy import openpyxl
        openpyxl, Font, get_column_letter = _import_openpyxl()
        
        logger.info(f"Preprocessing Excel file: {input_path}")
        wb = openpyxl.load_workbook(input_path)
        
        for sheet in wb.worksheets:
            # Save original row heights & widths
            original_row_heights = {
                row: sheet.row_dimensions[row].height 
                for row in range(1, sheet.max_row + 1)
            }
            original_col_widths = {
                col: sheet.column_dimensions[get_column_letter(col)].width 
                for col in range(1, sheet.max_column + 1)
            }

            # Insert a new row at the top for sheet title
            sheet.insert_rows(1)
            cell = sheet.cell(row=1, column=1)
            cell.value = sheet.title

            # Calculate the maximum length of content in column A
            max_length = max(
                len(sheet.title), 
                max(
                    (len(str(value)) for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True) 
                     for value in row if value is not None), 
                    default=0
                )
            )

            # Adjust column width based on the maximum length
            sheet.column_dimensions['A'].width = (max_length * 1.2) + 2

            # Set consistent font size
            cell.font = Font(size=10, name='Arial')

            # Set the height of the new row
            sheet.row_dimensions[1].height = 15

            # Restore original row heights (shifted by 1)
            for row, height in original_row_heights.items():
                sheet.row_dimensions[row + 1].height = height

            # Apply formatting to all columns
            for col in range(1, sheet.max_column + 1):
                # Convert cell values to formatted strings and set font
                for row in sheet.iter_rows(min_col=col, max_col=col):
                    for cell in row:
                        if cell.value is not None:
                            # Format numbers with 2 decimal places
                            if isinstance(cell.value, (int, float)):
                                cell.value = f"{cell.value:.2f}"
                            else:
                                cell.value = str(cell.value)
                            # Set consistent font
                            cell.font = Font(size=10, bold=False, name='Arial')

                # Calculate maximum length of values in the column
                max_length_col = max(
                    (len(str(value)) for row in sheet.iter_rows(min_col=col, max_col=col, values_only=True) 
                     for value in row if value is not None),
                    default=0
                )
                # Add buffer to width
                sheet.column_dimensions[get_column_letter(col)].width = (max_length_col * 1.2) + 3

            # Set page layout for better PDF output
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.print_options.horizontalCentered = True
            sheet.print_options.verticalCentered = True
            sheet.page_margins.left = 0.05
            sheet.page_margins.right = 0.05
            sheet.page_margins.top = 0.05
            sheet.page_margins.bottom = 0.05

            max_row = sheet.max_row
            max_col = sheet.max_column

            # Adjust scaling based on content size
            if max_row > 20 or max_col > 10:
                sheet.page_setup.fitToWidth = 1
                sheet.page_setup.fitToHeight = 1
                sheet.page_setup.scale = 50
            else:
                sheet.page_setup.fitToWidth = False
                sheet.page_setup.fitToHeight = False
                sheet.page_setup.scale = 100

            # Set print area
            sheet.print_area = f'A1:{get_column_letter(max_col)}{max_row}'

            # Set first row as repeating header
            sheet.print_title_rows = '1:1'

        # Save the modified workbook
        wb.save(output_path)
        logger.info(f"Excel preprocessing complete: {output_path}")
        return output_path
        
    except Exception as e:
        logger.error(f"Error preprocessing Excel file: {str(e)}")
        raise


def convert_to_pdf_libreoffice(input_path: str, output_dir: str, is_excel: bool = False) -> str:
    """
    Convert file to PDF using LibreOffice.
    
    Args:
        input_path: Path to input file
        output_dir: Directory where PDF should be saved
        is_excel: Whether the file is Excel (for special preprocessing)
        
    Returns:
        Path to the generated PDF file
        
    Raises:
        RuntimeError: If LibreOffice is not found or conversion fails
    """
    try:
        file_to_convert = input_path
        
        # Special preprocessing for Excel files
        if is_excel:
            logger.info("Applying Excel preprocessing")
            preprocessed_path = os.path.join(output_dir, "preprocessed_" + os.path.basename(input_path))
            file_to_convert = preprocess_excel(input_path, preprocessed_path)
        
        # Get LibreOffice path from config
        soffice_path = settings.libreoffice_path

        # Verify LibreOffice exists at configured path
        if not os.path.exists(soffice_path):
            raise RuntimeError(
                f"LibreOffice not found at configured path: {soffice_path}\n"
                "Please install LibreOffice or update LIBREOFFICE_PATH in .env:\n"
                "  macOS default: /Applications/LibreOffice.app/Contents/MacOS/soffice\n"
                "  Linux: /usr/bin/soffice\n"
                "  Docker: Set LIBREOFFICE_PATH environment variable"
            )

        logger.info(f"Using LibreOffice at: {soffice_path}")
        
        # Attempt conversion with retries
        max_attempts = 3
        for attempt in range(max_attempts):
            try:
                logger.info(f"Attempting LibreOffice conversion (attempt {attempt + 1}/{max_attempts})")
                logger.info(f"Converting: {file_to_convert}")
                
                # Run LibreOffice conversion
                process = subprocess.Popen(
                    [
                        soffice_path,  # Use configured path
                        "--headless",
                        "--convert-to",
                        "pdf",
                        os.path.basename(file_to_convert),
                        "--outdir",
                        output_dir
                    ],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    cwd=os.path.dirname(file_to_convert)
                )
                
                stdout, stderr = process.communicate(timeout=300)
                
                if process.returncode != 0:
                    error_msg = stderr.decode() if stderr else "Unknown error"
                    logger.error(f"LibreOffice conversion failed: {error_msg}")
                    logger.error(f"LibreOffice stdout: {stdout.decode() if stdout else 'No output'}")
                    raise RuntimeError(f"LibreOffice conversion failed: {error_msg}")
                
                logger.info(f"Files in output directory: {os.listdir(output_dir)}")
                
                # Find the generated PDF
                pdf_files = [f for f in os.listdir(output_dir) if f.endswith('.pdf')]
                if not pdf_files:
                    raise RuntimeError("No PDF file generated by LibreOffice")
                
                output_pdf_path = os.path.join(output_dir, pdf_files[0])
                
                if os.path.exists(output_pdf_path):
                    file_size = os.path.getsize(output_pdf_path)
                    logger.info(f"Successfully converted to PDF: {output_pdf_path} ({file_size} bytes)")
                    return output_pdf_path
                else:
                    raise RuntimeError(f"PDF file not found at expected path: {output_pdf_path}")
                    
            except subprocess.TimeoutExpired:
                process.kill()
                logger.error("LibreOffice conversion timed out after 300 seconds")
                if attempt < max_attempts - 1:
                    time.sleep(2)
                    continue
                raise RuntimeError("LibreOffice conversion timed out")
                
            except Exception as e:
                logger.error(f"Error during LibreOffice conversion: {str(e)}")
                if attempt < max_attempts - 1:
                    time.sleep(2)
                    continue
                raise
        
        raise RuntimeError("Failed to convert file to PDF after multiple attempts")
        
    except Exception as e:
        logger.error(f"Error in convert_to_pdf_libreoffice: {str(e)}")
        raise


def convert_file_to_pdf(input_file_content: bytes, filename: str) -> Tuple[bytes, str]:
    """
    Main conversion function: converts a file to PDF.
    
    Args:
        input_file_content: File content as bytes
        filename: Original filename
        
    Returns:
        Tuple of (PDF content as bytes, PDF filename)
        
    Raises:
        RuntimeError: If conversion fails
    """
    temp_dir = None
    
    try:
        # Create temporary directory
        temp_dir = create_temp_directory()
        
        # Save input file
        input_path = os.path.join(temp_dir, filename)
        with open(input_path, 'wb') as f:
            f.write(input_file_content)
        logger.info(f"Saved input file: {input_path} ({len(input_file_content)} bytes)")
        
        # Check if this is an Excel file for preprocessing
        is_excel = filename.lower().endswith(('.xls', '.xlsx'))
        
        # Convert to PDF using LibreOffice
        pdf_path = convert_to_pdf_libreoffice(input_path, temp_dir, is_excel=is_excel)
        
        # Read the generated PDF
        with open(pdf_path, 'rb') as f:
            pdf_content = f.read()
        
        # Generate PDF filename
        pdf_filename = os.path.splitext(filename)[0] + '.pdf'
        
        logger.info(f"Conversion complete: {pdf_filename} ({len(pdf_content)} bytes)")
        return pdf_content, pdf_filename
        
    except Exception as e:
        logger.error(f"Error converting file {filename}: {str(e)}")
        raise
        
    finally:
        # Clean up temporary files
        if temp_dir:
            cleanup_temp_files(temp_dir)

